import io
import os
import sys
import time
import base64
from datetime import datetime
from queue import Queue
from threading import Thread, Lock
from typing import Optional, Tuple, Dict
from openpyxl import Workbook, load_workbook
from pathlib import Path

import requests
from fastapi import FastAPI, Body, HTTPException
from fastapi.responses import JSONResponse
import uvicorn

from rich.table import Table
from rich.console import Console
from rich import box

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import uuid

# Add parent folder to sys.path for imports if needed
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)

# Add helper for service URLs: prefer env var, allow USE_LOCAL, otherwise use known remote defaults
REMOTE_DEFAULTS = {
    "UPLOADER_URL": "http://172.20.10.3:50060/upload",
    "NOISE_URL": "http://172.20.10.3:50061/reduce",
    "BACKGROUND_URL": "http://172.20.10.3:50062/remove",
    "CLASSIFIER_URL": "http://172.20.10.3:50063/classify",
}

def _get_service_url(env_name: str, local_default: str) -> str:
    val = os.getenv(env_name)
    if val:
        return val
    if os.getenv("USE_LOCAL", "false").lower() in ("1", "true", "yes"):
        print(f"[orchestrator] Warning: {env_name} not set — falling back to local default {local_default} (USE_LOCAL enabled)")
        return local_default
    # Fall back to remote defaults (containers on other physical machines)
    if env_name in REMOTE_DEFAULTS:
        remote = REMOTE_DEFAULTS[env_name]
        print(f"[orchestrator] Info: {env_name} not set — using remote default {remote}")
        return remote
    # Last resort: raise
    raise RuntimeError(f"Environment variable {env_name} is not set and no default is available.")

try:
    UPLOADER_URL = _get_service_url("UPLOADER_URL", "http://localhost:50060/upload")
    NOISE_URL = _get_service_url("NOISE_URL", "http://localhost:50061/reduce")
    BACKGROUND_URL = _get_service_url("BACKGROUND_URL", "http://localhost:50062/remove")
    CLASSIFIER_URL = _get_service_url("CLASSIFIER_URL", "http://localhost:50063/classify")
except RuntimeError as e:
    print(f"[orchestrator][startup error] {e}")
    raise

print("UPLOADER_URL =", UPLOADER_URL)
print("NOISE_URL =", NOISE_URL)
print("BACKGROUND_URL =", BACKGROUND_URL)
print("CLASSIFIER_URL =", CLASSIFIER_URL)

EXCEL_PATH = "/app/results/pipeline_results.xlsx"
OUTPUT_DIR = "/app/results"


# # Environment variables for downstream REST service URLs
# UPLOADER_URL = os.getenv("UPLOADER_URL", "http://localhost:50060/upload")
# NOISE_URL = os.getenv("NOISE_URL", "http://localhost:50061/reduce")
# BACKGROUND_URL = os.getenv("BACKGROUND_URL", "http://localhost:50062/remove")
# CLASSIFIER_URL = os.getenv("CLASSIFIER_URL", "http://localhost:50063/classify")

app = FastAPI(title="Image Processor Orchestrator REST")

# ===== Helper functions =====
def format_timestamp(ts: Optional[float]) -> str:
    if ts is None:
        return "N/A"
    return datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

def _b64encode(data: bytes) -> str:
    return base64.b64encode(data).decode("utf-8")

def _b64decode(data: str) -> bytes:
    return base64.b64decode(data.encode("utf-8"))

# ===== Orchestrator =====
class OrchestratorService:
    def __init__(self):
        self.timeline_data = []
        self.timeline_lock = Lock()
        self.task_results: Dict[str, dict] = {}  # store task results

        # Queues for pipeline stages
        self.uploader_queue = Queue()
        self.noise_queue = Queue()
        self.bg_queue = Queue()
        self.clf_queue = Queue()

        # Start worker threads for each stage
        Thread(target=self.stage_worker, args=(self.uploader_queue, self.noise_queue, "Uploader", UPLOADER_URL), daemon=True).start()
        Thread(target=self.stage_worker, args=(self.noise_queue, self.bg_queue, "Noise Reduction", NOISE_URL), daemon=True).start()
        Thread(target=self.stage_worker, args=(self.bg_queue, self.clf_queue, "Background Removal", BACKGROUND_URL), daemon=True).start()
        Thread(target=self.stage_worker, args=(self.clf_queue, None, "Classifier", CLASSIFIER_URL), daemon=True).start()

    def stage_worker(self, input_queue: Queue, output_queue: Optional[Queue], stage_name: str, service_url: str):
        """Worker thread for a pipeline stage (REST version)."""
        while True:
            item = input_queue.get()
            if item is None:
                break
            image_data, file_name, task_id, resp_future = item

            # Initialize timeline entry
            timeline = resp_future.setdefault("timeline", {})
            if stage_name not in timeline:
                timeline[stage_name] = {"start": None, "end": None, "events": []}

            try:
                # Record start
                start_time = time.time()
                timeline[stage_name]["start"] = start_time
                timeline[stage_name]["events"].append(("start", start_time))

                # Prepare payload
                payload = {"image_data": _b64encode(image_data)}
                if stage_name == "Uploader":
                    payload["file_name"] = file_name

                # Call downstream REST service with retries
                result = None
                max_attempts = 3
                backoff = 1.0
                for attempt in range(1, max_attempts + 1):
                    try:
                        print(f"[orchestrator] Calling {service_url} for stage '{stage_name}' file '{file_name}' (attempt {attempt})")
                        resp = requests.post(service_url, json=payload, timeout=30)
                        resp.raise_for_status()
                        result = resp.json()
                        break
                    except requests.exceptions.RequestException as re:
                        print(f"[orchestrator][warning] attempt {attempt} failed for {service_url}: {re}")
                        if attempt < max_attempts:
                            time.sleep(backoff)
                            backoff *= 2
                        else:
                            # Permanent failure -> record error and timeline end
                            err_time = time.time()
                            timeline[stage_name]["end"] = err_time
                            timeline[stage_name]["events"].append(("error", err_time))
                            resp_future["error"] = f"Request to {service_url} failed after {max_attempts} attempts: {re}"
                            print(f"[orchestrator][error] {resp_future['error']}")
                            result = None

                if result is None:
                    # stop processing this task further
                    continue

                # Record end
                end_time = time.time()
                timeline[stage_name]["end"] = end_time
                timeline[stage_name]["events"].append(("end", end_time))

                # Get next image data
                next_image_data = _b64decode(result.get("image_data", "")) if result.get("image_data") else image_data

                if output_queue:
                    # Put image into next stage
                    output_queue.put((next_image_data, file_name, task_id, resp_future))
                else:
                    # Last stage -> store results
                    resp_future["response"] = result
                    resp_future["final_image_data"] = next_image_data
                    resp_future["completed_time"] = time.time()

                    # Append to timeline_data for plotting
                    with self.timeline_lock:
                        self.timeline_data.append({
                            "image_name": file_name,
                            "enqueue_time": resp_future["enqueue_time"],
                            "timeline": {s: {"start": t["start"], "end": t["end"], "events": t.get("events", [])} 
                                        for s, t in resp_future["timeline"].items()}
                        })

            except Exception as e:
                # Ensure timeline end recorded on unexpected exceptions
                err_time = time.time()
                try:
                    timeline[stage_name]["end"] = err_time
                    timeline[stage_name]["events"].append(("error", err_time))
                except Exception:
                    pass
                resp_future["error"] = f"{e}"
                print(f"[orchestrator][exception] Stage '{stage_name}' for '{file_name}' failed: {e}")
            finally:
                input_queue.task_done()

    def _plot_timeline(self, output_file: str | None = None) -> tuple[bytes, str]:
        """Generate a timeline Gantt chart."""
        with self.timeline_lock:
            if not self.timeline_data:
                return b"", ""
            timeline_data_copy = []
            for img_data in self.timeline_data:
                timeline_copy = {s: {"start": t["start"], "end": t["end"]} for s, t in img_data["timeline"].items()}
                timeline_data_copy.append({
                    "image_name": img_data["image_name"],
                    "enqueue_time": img_data["enqueue_time"],
                    "timeline": timeline_copy
                })

        services = ["Uploader", "Noise Reduction", "Background Removal", "Classifier"]
        fig, ax = plt.subplots(figsize=(14, 8))
        colors = plt.cm.Set3(range(len(timeline_data_copy)))

        y_positions = {service: i for i, service in enumerate(services)}
        first_enqueue = min(d["enqueue_time"] for d in timeline_data_copy)

        for img_idx, img_data in enumerate(timeline_data_copy):
            image_name = img_data["image_name"]
            timeline = img_data["timeline"]

            for service in services:
                if service in timeline and timeline[service]["start"] is not None and timeline[service]["end"] is not None:
                    start_time = timeline[service]["start"] - first_enqueue
                    end_time = timeline[service]["end"] - first_enqueue
                    duration = end_time - start_time
                    if duration <= 0:
                        continue
                    ax.barh(
                        y_positions[service],
                        duration,
                        left=start_time,
                        height=0.6,
                        color=colors[img_idx],
                        alpha=0.7,
                        edgecolor="black",
                        linewidth=0.5,
                        label=image_name if service == services[0] else ""
                    )

        ax.set_yticks([y_positions[s] for s in services])
        ax.set_yticklabels(services)
        ax.set_xlabel("Time (seconds)", fontsize=12, fontweight="bold")
        ax.set_ylabel("Service", fontsize=12, fontweight="bold")
        ax.set_title("Pipeline Service Execution Timeline", fontsize=14, fontweight="bold", pad=20)
        ax.grid(True, alpha=0.3, axis="x")
        ax.set_axisbelow(True)
        ax.set_xlim(left=0)

        handles, labels = ax.get_legend_handles_labels()
        if handles:
            by_label = dict(zip(labels, handles))
            ax.legend(by_label.values(), by_label.keys(), loc="upper left", bbox_to_anchor=(1.02, 1), title="Images", fontsize=9)

        plt.tight_layout()
        buffer = io.BytesIO()
        plt.savefig(buffer, format="png", dpi=150, bbox_inches="tight")
        buffer.seek(0)
        image_bytes = buffer.getvalue()
        buffer.close()

        if output_file is None:
            output_file = os.path.join(OUTPUT_DIR, "pipeline_timeline.png")
        with open(output_file, "wb") as f:
            f.write(image_bytes)

        plt.close(fig)
        return image_bytes, output_file

# Create global orchestrator instance
orchestrator = OrchestratorService()

# ===== REST endpoints =====
@app.post("/process")
async def process_image(payload: dict = Body(...)):
    img_b64 = payload.get("image_data")
    name = payload.get("file_name") or "unknown"
    if not img_b64:
        return JSONResponse(status_code=400, content={
            "success": False,
            "message": "No file or image_data provided"
        })

    raw = _b64decode(img_b64)
    start_time = time.time()
    task_id = str(uuid.uuid4())
    resp_future = {"enqueue_time": start_time, "file_name": name}

    # Store the future in orchestrator for later status checking
    orchestrator.task_results[task_id] = resp_future

    # Enqueue the image for pipeline processing
    orchestrator.uploader_queue.put((raw, name, task_id, resp_future))

    # Return task accepted immediately
    return JSONResponse(content={
        "success": True,
        "message": f"Image '{name}' accepted for processing.",
        "task_id": task_id
    })

@app.get("/status/{task_id}")
async def get_status(task_id: str):
    resp_future = orchestrator.task_results.get(task_id)
    if not resp_future:
        raise HTTPException(status_code=404, detail="Task not found")

    file_name = resp_future.get("file_name", "unknown")

    # Check if all stages for this image are completed
    timeline = resp_future.get("timeline", {})
    all_stages_done = all(info.get("start") is not None and info.get("end") is not None
                          for info in timeline.values())

    if "response" in resp_future or "error" in resp_future:
        console = Console()

        # Completed → save timeline and print table
        if all_stages_done and timeline:

            def save_timeline_to_excel(file_name, timeline):
                # Check if file exists — if yes, load it; otherwise create new workbook
                if Path(EXCEL_PATH).exists():
                    wb = load_workbook(EXCEL_PATH)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Results"
                    ws.append(["Stage / Image"])  # First cell header

                # Check if the image column already exists
                headers = [cell.value for cell in ws[1]]
                if file_name not in headers:
                    ws.cell(row=1, column=len(headers) + 1, value=file_name)

                image_col = headers.index(file_name) + 1 if file_name in headers else ws.max_column

                # Insert or update rows for each processing stage
                for stage, info in timeline.items():
                    # Find row for this stage, or create a new one
                    row_num = None
                    for row in range(2, ws.max_row + 1):
                        if ws.cell(row=row, column=1).value == stage:
                            row_num = row
                            break

                    if row_num is None:
                        row_num = ws.max_row + 1
                        ws.cell(row=row_num, column=1, value=stage)

                    processing_time_ms = (info["end"] - info["start"]) * 1000
                    ws.cell(row=row_num, column=image_col, value=processing_time_ms)

                wb.save(EXCEL_PATH)
                print(f"📁 Excel updated: {EXCEL_PATH}")

            save_timeline_to_excel(file_name, resp_future["timeline"])

            table = Table(
                title=f"📊 REST Pipeline Performance Summary - {file_name}",
                show_header=True,
                header_style="bold magenta",
                box=box.ROUNDED
            )
            table.add_column("Stage", justify="center", style="cyan", no_wrap=True)
            table.add_column("Start Timestamp", justify="center", style="green")
            table.add_column("End Timestamp", justify="center", style="green")
            table.add_column("Processing Time (ms)", justify="center", style="yellow")

            for stage, info in resp_future["timeline"].items():
                start_ts = format_timestamp(info["start"])
                end_ts = format_timestamp(info["end"])
                processing_time_ms = (info["end"] - info["start"]) * 1000
                table.add_row(stage, start_ts, end_ts, f"{processing_time_ms:.3f}")
            console.print(table)

        plot_bytes, plot_path = orchestrator._plot_timeline()
        return {
            "success": "response" in resp_future,
            "message": "Processing completed",
            "response": resp_future.get("response"),
            "processed_image_data": _b64encode(resp_future.get("final_image_data", b"")),
            "timeline_plot_image": _b64encode(plot_bytes) if plot_bytes else "",
            "timeline_plot_filename": plot_path if plot_bytes else "",
            "error": resp_future.get("error")
        }
    else:
        # Pending
        return {"success": False, "message": "Processing pending"}

# ===== Serve =====
def serve(port: int = 50051):
    uvicorn.run(app, host="0.0.0.0", port=port)

if __name__ == "__main__":
    serve()
