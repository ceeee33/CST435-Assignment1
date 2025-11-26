import io
import os
import sys
import time
import grpc
from PIL import Image
from datetime import datetime
from queue import Queue
from threading import Thread, Lock
from typing import Optional, Tuple
from concurrent import futures
from openpyxl import Workbook, load_workbook
from pathlib import Path

from rich.table import Table
from rich.console import Console
from rich import box

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt

current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)

import pipeline_pb2
import pipeline_pb2_grpc

# Environment variables for service addresses
UPLOADER_ADDR = os.getenv("UPLOADER_ADDR", "192.168.68.122:50040")
NOISE_ADDR = os.getenv("NOISE_ADDR", "192.168.68.122:50041")
BACKGROUND_ADDR = os.getenv("BACKGROUND_ADDR", "192.168.68.122:50042")
CLASSIFIER_ADDR = os.getenv("CLASSIFIER_ADDR", "192.168.68.122:50043")

EXCEL_PATH = "/app/results/pipeline_results.xlsx"
OUTPUT_DIR = "/app/results"

def format_timestamp(ts: Optional[float]) -> str:
    if ts is None:
        return "N/A"
    return datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]


class OrchestratorServicer(pipeline_pb2_grpc.ImageProcessorOrchestratorServicer):
    def __init__(self):
        # Connect to downstream services
        self.uploader_channel = grpc.insecure_channel(UPLOADER_ADDR)
        self.noise_channel = grpc.insecure_channel(NOISE_ADDR)
        self.bg_channel = grpc.insecure_channel(BACKGROUND_ADDR)
        self.clf_channel = grpc.insecure_channel(CLASSIFIER_ADDR)

        self.uploader = pipeline_pb2_grpc.ImageUploaderStub(self.uploader_channel)
        self.noise = pipeline_pb2_grpc.NoiseReducerStub(self.noise_channel)
        self.bg = pipeline_pb2_grpc.BackgroundRemoverStub(self.bg_channel)
        self.clf = pipeline_pb2_grpc.ImageClassifierStub(self.clf_channel)
        
        # Store timeline data for plotting (thread-safe)
        self.timeline_data = []  # List of dicts: {image_name, enqueue_time, timeline}
        self.timeline_lock = Lock()

        # Wait for channels to be ready
        for name, ch, addr in [
            ("uploader", self.uploader_channel, UPLOADER_ADDR),
            ("noise", self.noise_channel, NOISE_ADDR),
            ("background", self.bg_channel, BACKGROUND_ADDR),
            ("classifier", self.clf_channel, CLASSIFIER_ADDR),
        ]:
            try:
                print(f"[orchestrator] Waiting for {name} at {addr}...")
                grpc.channel_ready_future(ch).result(timeout=120)
                print(f"[orchestrator] {name} channel ready")
            except Exception as e:
                print(f"[orchestrator] {name} not ready after timeout: {e}")

        # Queues for pipeline stages
        self.uploader_queue = Queue()
        self.noise_queue = Queue()
        self.bg_queue = Queue()
        self.clf_queue = Queue()

        # Start worker threads
        Thread(target=self.stage_worker, args=(self.uploader_queue, self.noise_queue, self.uploader, "Uploader"), daemon=True).start()
        Thread(target=self.stage_worker, args=(self.noise_queue, self.bg_queue, self.noise, "Noise Reduction"), daemon=True).start()
        Thread(target=self.stage_worker, args=(self.bg_queue, self.clf_queue, self.bg, "Background Removal"), daemon=True).start()
        Thread(target=self.stage_worker, args=(self.clf_queue, None, self.clf, "Classifier"), daemon=True).start()

    def stage_worker(self, input_queue: Queue, output_queue: Queue, stub, rpc_name: str):
        """Worker thread for a pipeline stage."""
        while True:
            item = input_queue.get()
            if item is None:  # shutdown signal
                break
            image_data, file_name, resp_future = item

            # # Initialize timeline dict for this image
            timeline = resp_future.setdefault("timeline", {})
            # timeline[rpc_name] = {"start": time.time()}  # record start
            # Initialize entry for this RPC if not exists
            if rpc_name not in timeline:
                timeline[rpc_name] = {"start": None, "end": None, "events": []}
            
            try:                
                # Record start time
                start_time = time.time()
                timeline[rpc_name]["start"] = start_time
                timeline[rpc_name]["events"].append(("start", start_time))

                # Determine the RPC call and request type
                if rpc_name == "Uploader":
                    rpc_call = stub.Upload
                    request = pipeline_pb2.UploadRequest(image_data=image_data, file_name=file_name)
                elif rpc_name == "Noise Reduction":
                    rpc_call = stub.Reduce
                    request = pipeline_pb2.NoiseReductionRequest(image_data=image_data)
                elif rpc_name == "Background Removal":
                    rpc_call = stub.Remove
                    request = pipeline_pb2.BackgroundRemoveRequest(image_data=image_data)
                elif rpc_name == "Classifier":
                    rpc_call = stub.Classify
                    request = pipeline_pb2.ClassifyRequest(image_data=image_data)
                else:
                    raise ValueError(f"Unknown RPC name: {rpc_name}")

                # Call the service
                result = rpc_call(request)

                # Record end time
                end_time = time.time()
                timeline[rpc_name]["end"] = end_time
                timeline[rpc_name]["events"].append(("end", end_time))

                # Pass to next stage or complete the response
                next_image_data = getattr(result, 'image_data', image_data)
                
                if output_queue:
                    output_queue.put((next_image_data, file_name, resp_future))
                else:
                    # Last stage (classifier)
                    resp_future["response"] = result
                    resp_future["final_image_data"] = image_data

            except Exception as e:
                resp_future["error"] = str(e)
            finally:
                input_queue.task_done()

    def _plot_timeline(self, output_file: Optional[str] = None) -> Tuple[bytes, str]:
        """Generate a Gantt chart showing service execution timeline for all processed images.

        Returns:
            A tuple containing the PNG bytes of the plot and the file path where it was saved.
        """
        # Make a local copy of timeline data for thread safety
        with self.timeline_lock:
            if not self.timeline_data:
                print("[Plot] No timeline data available to plot.")
                return b"", ""
            # Deep copy the data to avoid issues during plotting
            timeline_data_copy = []
            for img_data in self.timeline_data:
                timeline_copy = {}
                for service, info in img_data["timeline"].items():
                    timeline_copy[service] = {
                        "start": info["start"],
                        "end": info["end"]
                    }
                timeline_data_copy.append({
                    "image_name": img_data["image_name"],
                    "enqueue_time": img_data["enqueue_time"],
                    "timeline": timeline_copy
                })

        # Prepare data for plotting
        services = ["Uploader", "Noise Reduction", "Background Removal", "Classifier"]

        fig, ax = plt.subplots(figsize=(14, 8))

        # Color map for different images
        colors = plt.cm.Set3(range(len(timeline_data_copy)))

        y_positions = {}
        y_offset = 0

        # Plot each service as a row
        for service in services:
            y_positions[service] = y_offset
            y_offset += 1

        # Find the earliest timestamp to normalize all times
        first_enqueue = min(d["enqueue_time"] for d in timeline_data_copy)

        # Plot bars for each image's service execution
        for img_idx, img_data in enumerate(timeline_data_copy):
            image_name = img_data["image_name"]
            timeline = img_data["timeline"]

            for service in services:
                if service in timeline and timeline[service]["start"] is not None:
                    # Normalize timestamps relative to first enqueue
                    start_time = timeline[service]["start"] - first_enqueue
                    end_time = timeline[service]["end"] - first_enqueue
                    duration = end_time - start_time

                    if duration <= 0:
                        continue

                    y_pos = y_positions[service]

                    # Plot horizontal bar
                    ax.barh(
                        y_pos,
                        duration,
                        left=start_time,
                        height=0.6,
                        color=colors[img_idx],
                        alpha=0.7,
                        edgecolor="black",
                        linewidth=0.5,
                        label=image_name if service == services[0] else ""
                    )

                    # # Add text label on the bar (if space allows)
                    # if duration > 0.01:  # Only label if bar is wide enough
                    #     mid_time = start_time + duration / 2
                    #     ax.text(
                    #         mid_time,
                    #         y_pos,
                    #         image_name[:15],  # Truncate long names
                    #         ha="center",
                    #         va="center",
                    #         fontsize=7,
                    #         fontweight="bold",
                    #         color="white" if img_idx % 2 == 0 else "black"
                    #     )

        # Formatting
        ax.set_yticks([y_positions[s] for s in services])
        ax.set_yticklabels(services)
        ax.set_xlabel("Time (seconds)", fontsize=12, fontweight="bold")
        ax.set_ylabel("Service", fontsize=12, fontweight="bold")
        ax.set_title("Pipeline Service Execution Timeline", fontsize=14, fontweight="bold", pad=20)
        ax.grid(True, alpha=0.3, axis="x")
        ax.set_axisbelow(True)
        ax.set_xlim(left=0)

        # Add legend (one entry per image)
        handles, labels = ax.get_legend_handles_labels()
        if handles:
            # Remove duplicates
            by_label = dict(zip(labels, handles))
            ax.legend(
                by_label.values(),
                by_label.keys(),
                loc="upper left",
                bbox_to_anchor=(1.02, 1),
                title="Images",
                fontsize=9,
            )

        plt.tight_layout()

        buffer = io.BytesIO()
        plt.savefig(buffer, format="png", dpi=150, bbox_inches="tight")
        buffer.seek(0)
        image_bytes = buffer.getvalue()
        buffer.close()

        if output_file is None:
            output_file = os.path.join(OUTPUT_DIR, "pipeline_timeline.png")

        with open(output_file, "wb") as f:
            f.write(image_bytes)

        plt.close(fig)

        print(f"[Plot] Timeline graph saved to {output_file}")
        return image_bytes, output_file

    def Process(self, request: pipeline_pb2.ProcessRequest, context) -> pipeline_pb2.ProcessResponse:
        """Enqueue the image and wait for pipeline to finish."""
        start_time = time.time()
        resp_future = {}
        plot_bytes: bytes = b""
        plot_path: str = ""

        # Add enqueue time
        resp_future["enqueue_time"] = start_time

        # Enqueue image into the first stage
        self.uploader_queue.put((request.image_data, request.file_name, resp_future))

        # Wait for final response from classifier
        while "response" not in resp_future and "error" not in resp_future:
            time.sleep(0.01)  # small sleep to avoid busy-waiting
        
        def save_timeline_to_excel(file_name, timeline):
            # Check if file exists — if yes, load it; otherwise create new workbook
            if Path(EXCEL_PATH).exists():
                wb = load_workbook(EXCEL_PATH)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Results"
                ws.append(["Stage / Image"])  # First cell header

            # Check if the image column already exists
            headers = [cell.value for cell in ws[1]]
            if file_name not in headers:
                ws.cell(row=1, column=len(headers) + 1, value=file_name)

            image_col = headers.index(file_name) + 1 if file_name in headers else ws.max_column

            # Insert or update rows for each processing stage
            for stage, info in timeline.items():
                # Find row for this stage, or create a new one
                row_num = None
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value == stage:
                        row_num = row
                        break

                if row_num is None:
                    row_num = ws.max_row + 1
                    ws.cell(row=row_num, column=1, value=stage)

                processing_time_ms = (info["end"] - info["start"]) * 1000
                ws.cell(row=row_num, column=image_col, value=processing_time_ms)

            wb.save(EXCEL_PATH)
            print(f"📁 Excel updated: {EXCEL_PATH}")

        save_timeline_to_excel(request.file_name, resp_future["timeline"])

        # --------------------------
        # table summary
        # --------------------------
        table = Table(title=f"📊 gRPC Pipeline Performance Summary - {request.file_name}",
                    show_header=True, header_style="bold magenta", box=box.ROUNDED)
        table.add_column("Stage", justify="center", style="cyan", no_wrap=True)
        table.add_column("Start Timestamp", justify="center", style="green")
        table.add_column("End Timestamp", justify="center", style="green")
        table.add_column("Processing Time (ms)", justify="center", style="yellow")

        # Print timeline for debugging
        if "timeline" in resp_future:
            print(f"\n[Timeline] Image '{request.file_name}':")
            for stage, info in resp_future["timeline"].items():
                start_ts = format_timestamp(info["start"])
                end_ts = format_timestamp(info["end"])
                processing_time_ms = (info["end"] - info["start"]) * 1000
                print(f"  {stage} Start at {start_ts},  ({info['start'] - start_time:.3f}s from enqueue)")
                print(f"  {stage} End at {end_ts}, ({info['end'] - start_time:.3f}s from enqueue)")
                
                # Add row to table (without curly braces - they create sets!)
                table.add_row(stage, start_ts, end_ts, f"{processing_time_ms:.3f}")
            
            # Print the table using Rich Console
            console = Console()
            console.print(table)
            
            # Store timeline data for plotting (thread-safe)
            with self.timeline_lock:
                # Deep copy timeline to avoid reference issues
                timeline_copy = {}
                for service, info in resp_future["timeline"].items():
                    timeline_copy[service] = {
                        "start": info["start"],
                        "end": info["end"],
                        "events": info.get("events", []).copy() if "events" in info else []
                    }
                
                self.timeline_data.append({
                    "image_name": request.file_name,
                    "enqueue_time": start_time,
                    "timeline": timeline_copy
                })
            
            # Generate plot after each image (outside lock to avoid blocking)
            try:
                plot_bytes, plot_path = self._plot_timeline()
            except Exception as e:
                print(f"[Plot] Error generating plot: {e}")

        if "error" in resp_future:
            return pipeline_pb2.ProcessResponse(
                processed_image_data=b"",
                classification_label="",
                success=False,
                message=resp_future["error"],
                response_time_ms=0.0
            )

        # Build the response
        cls_result = resp_future["response"]
        total_time_ms = (time.time() - start_time) * 1000
        return pipeline_pb2.ProcessResponse(
            processed_image_data=resp_future.get("final_image_data", b""),
            classification_label=getattr(cls_result, 'label', ''),
            success=True,
            message=f"Image processed successfully! total={total_time_ms:.2f}ms",
            response_time_ms=total_time_ms,
            timeline_plot_image=plot_bytes,
            timeline_plot_filename=plot_path
        )


def serve(port: int = 50030):
    """Start the gRPC orchestrator server."""
    server = grpc.server(futures.ThreadPoolExecutor(max_workers=10))
    pipeline_pb2_grpc.add_ImageProcessorOrchestratorServicer_to_server(OrchestratorServicer(), server)
    server.add_insecure_port(f"[::]:{port}")
    server.start()
    print(f"Orchestrator service started on port {port}")
    server.wait_for_termination()


if __name__ == "__main__":
    serve()
